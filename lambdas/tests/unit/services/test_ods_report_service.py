import os
import tempfile
from datetime import datetime, timezone
from unittest.mock import call

import pytest
from freezegun import freeze_time
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader

from enums.document_review_reason import DocumentReviewReason
from enums.document_review_status import DocumentReviewStatus
from enums.dynamo_filter import AttributeOperator
from enums.file_type import FileType
from enums.lambda_error import LambdaError
from enums.metadata_field_names import DocumentReferenceMetadataFields
from enums.patient_ods_inactive_status import PatientOdsInactiveStatus
from models.document_review import (
    DocumentReviewFileDetails,
    DocumentUploadReviewReference,
)
from services.ods_report_service import OdsReportService
from utils.common_query_filters import FinalStatusAndNotSuperseded
from utils.dynamo_query_filter_builder import DynamoQueryFilterBuilder
from utils.lambda_exceptions import OdsReportException


@pytest.fixture
def ods_report_service(mocker, set_env):
    mocker.patch("services.ods_report_service.S3Service")
    mocker.patch("services.ods_report_service.DynamoDBService")
    mocker.patch("services.ods_report_service.DocumentUploadReviewService")
    temp_folder = tempfile.mkdtemp()
    mocker.patch.object(tempfile, "mkdtemp", return_value=temp_folder)
    service = OdsReportService()
    return service


@pytest.fixture
def mocked_context(mocker):
    mocked_context = mocker.MagicMock()
    mocked_context.authorization = {
        "selected_organisation": {"org_ods_code": "Y12345"},
        "repository_role": "GP_ADMIN",
    }
    yield mocker.patch("services.ods_report_service.request_context", mocked_context)


@pytest.fixture
def mocked_pcse_context(mocker):
    mocked_pcse_context = mocker.MagicMock()
    mocked_pcse_context.authorization = {
        "selected_organisation": {"org_ods_code": "Y12345"},
        "repository_role": "PCSE",
    }
    yield mocker.patch(
        "services.ods_report_service.request_context",
        mocked_pcse_context,
    )


@pytest.fixture
def mock_review_result():
    return DocumentUploadReviewReference(
        nhs_number="mock_nhs_number",
        author="mock_author",
        review_reason=DocumentReviewReason.UNSUCCESSFUL_UPLOAD,
        document_snomed_code_type="mock_snomed_code",
        upload_date=int(datetime.now().timestamp()),
        files=[
            DocumentReviewFileDetails(
                file_name="mock_file_name",
                file_location="mock_file_location",
            ),
        ],
        custodian="mock_custodian",
    )


@pytest.fixture
def mock_create_and_save_ods_report(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service, "create_and_save_ods_report")


@pytest.fixture
def mock_scan_table_with_filter(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service, "scan_table_with_filter")


@pytest.fixture
def mock_query_table_by_index(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service, "query_table_by_index")


@pytest.fixture
def mock_dynamo_service_scan_table(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service.dynamo_service, "scan_whole_table")


@pytest.fixture
def mock_create_report_csv(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service, "create_csv_report")


@pytest.fixture
def mock_create_report_pdf(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service, "create_pdf_report")


@pytest.fixture
def mock_create_report_xlsx(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service, "create_xlsx_report")


@pytest.fixture
def mock_save_report_to_s3(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service, "save_report_to_s3")


@pytest.fixture
def mock_get_pre_signed_url(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service, "get_pre_signed_url")


@pytest.fixture
def mock_create_review_csv_report(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service, "create_review_csv_report")


def test_get_nhs_numbers_by_ods(
    ods_report_service,
    mock_query_table_by_index,
    mock_create_and_save_ods_report,
    mocker,
):
    mock_query_table_by_index.return_value = [
        {DocumentReferenceMetadataFields.NHS_NUMBER.value: "NHS123"},
        {DocumentReferenceMetadataFields.NHS_NUMBER.value: "NHS456"},
    ]

    patient_rows = {
        "NHS123": {"nhs_number": "NHS123", "created": None, "last_updated": None},
        "NHS456": {"nhs_number": "NHS456", "created": None, "last_updated": None},
    }
    mocker.patch.object(
        ods_report_service,
        "build_patient_rows",
        return_value=patient_rows,
    )

    ods_report_service.generate_ods_report("ODS123")

    mock_query_table_by_index.assert_called_once_with("ODS123")
    mock_create_and_save_ods_report.assert_called_once_with(
        ods_code="ODS123",
        patient_rows=patient_rows,
        create_pre_signed_url=False,
        upload_to_s3=False,
        file_type_output=FileType.CSV,
    )


def test_get_nhs_numbers_by_ods_with_temp_folder(
    ods_report_service,
    mock_query_table_by_index,
    mock_create_and_save_ods_report,
    mocker,
):
    mock_query_table_by_index.return_value = [
        {DocumentReferenceMetadataFields.NHS_NUMBER.value: "NHS123"},
        {DocumentReferenceMetadataFields.NHS_NUMBER.value: "NHS456"},
    ]

    patient_rows = {
        "NHS123": {"nhs_number": "NHS123", "created": None, "last_updated": None},
        "NHS456": {"nhs_number": "NHS456", "created": None, "last_updated": None},
    }
    mocker.patch.object(
        ods_report_service,
        "build_patient_rows",
        return_value=patient_rows,
    )

    ods_report_service.generate_ods_report("ODS123", is_upload_to_s3_needed=True)

    mock_query_table_by_index.assert_called_once_with("ODS123")
    mock_create_and_save_ods_report.assert_called_once_with(
        ods_code="ODS123",
        patient_rows=patient_rows,
        create_pre_signed_url=False,
        upload_to_s3=True,
        file_type_output=FileType.CSV,
    )
    assert ods_report_service.temp_output_dir != ""


def test_scan_table_with_filter(
    ods_report_service,
    mocked_context,
    mock_dynamo_service_scan_table,
):
    mock_dynamo_service_scan_table.return_value = [
        {DocumentReferenceMetadataFields.NHS_NUMBER.value: "NHS123"},
        {DocumentReferenceMetadataFields.NHS_NUMBER.value: "NHS456"},
        {DocumentReferenceMetadataFields.NHS_NUMBER.value: "NHS789"},
    ]

    results = ods_report_service.scan_table_with_filter("ODS123")

    assert len(results) == 3
    assert mock_dynamo_service_scan_table.call_count == 1


def test_scan_table_with_filter_no_results(
    ods_report_service,
    mocked_context,
    mock_dynamo_service_scan_table,
):
    mock_dynamo_service_scan_table.return_value = []

    with pytest.raises(OdsReportException):
        ods_report_service.scan_table_with_filter("ODS123")


@freeze_time("2024-01-01T12:00:00Z")
def test_create_and_save_ods_report_create_csv(
    ods_report_service,
    mock_create_report_csv,
    mock_save_report_to_s3,
    mock_get_pre_signed_url,
    mocker,
):
    ods_code = "ODS123"
    patient_rows = {
        "NHS123": {"nhs_number": "NHS123", "created": None, "last_updated": None},
        "NHS456": {"nhs_number": "NHS456", "created": None, "last_updated": None},
    }
    file_name = "LloydGeorgeSummary_ODS123_2_2024-01-01_12-00.csv"
    temp_file_path = os.path.join(ods_report_service.temp_output_dir, file_name)

    mocker.patch.object(
        ods_report_service,
        "create_ods_report",
        return_value=(file_name, temp_file_path),
    )

    result = ods_report_service.create_and_save_ods_report(
        ods_code,
        patient_rows,
        upload_to_s3=True,
        file_type_output=FileType.CSV,
    )

    mock_save_report_to_s3.assert_called_once_with(ods_code, file_name, temp_file_path)
    mock_get_pre_signed_url.assert_not_called()
    assert result is None


@freeze_time("2024-01-01T12:00:00Z")
def test_create_and_save_ods_report_create_pdf(
    ods_report_service,
    mock_create_report_pdf,
    mock_save_report_to_s3,
    mock_get_pre_signed_url,
    mocker,
):
    ods_code = "ODS123"
    patient_rows = {
        "NHS123": {"nhs_number": "NHS123", "created": None, "last_updated": None},
        "NHS456": {"nhs_number": "NHS456", "created": None, "last_updated": None},
    }
    file_name = "LloydGeorgeSummary_ODS123_2_2024-01-01_12-00.pdf"
    temp_file_path = os.path.join(ods_report_service.temp_output_dir, file_name)

    mocker.patch.object(
        ods_report_service,
        "create_ods_report",
        return_value=(file_name, temp_file_path),
    )

    ods_report_service.create_and_save_ods_report(
        ods_code,
        patient_rows,
        upload_to_s3=True,
        file_type_output=FileType.PDF,
    )

    mock_save_report_to_s3.assert_called_once_with(ods_code, file_name, temp_file_path)
    mock_get_pre_signed_url.assert_not_called()


@freeze_time("2024-01-01T12:00:00Z")
def test_create_and_save_ods_report_create_xlsx(
    ods_report_service,
    mock_create_report_xlsx,
    mock_save_report_to_s3,
    mock_get_pre_signed_url,
    mocker,
):
    ods_code = "ODS123"
    patient_rows = {
        "NHS123": {"nhs_number": "NHS123", "created": None, "last_updated": None},
        "NHS456": {"nhs_number": "NHS456", "created": None, "last_updated": None},
    }
    file_name = "LloydGeorgeSummary_ODS123_2_2024-01-01_12-00.xlsx"
    temp_file_path = os.path.join(ods_report_service.temp_output_dir, file_name)

    mocker.patch.object(
        ods_report_service,
        "create_ods_report",
        return_value=(file_name, temp_file_path),
    )

    ods_report_service.create_and_save_ods_report(
        ods_code,
        patient_rows,
        upload_to_s3=True,
        file_type_output=FileType.XLSX,
    )

    mock_save_report_to_s3.assert_called_once_with(ods_code, file_name, temp_file_path)
    mock_get_pre_signed_url.assert_not_called()


def test_create_and_save_ods_report_send_invalid_file_type(
    ods_report_service,
    mock_create_report_pdf,
    mock_save_report_to_s3,
    mock_get_pre_signed_url,
):
    ods_code = "ODS123"
    patient_rows = {
        "NHS123": {"nhs_number": "NHS123", "created": None, "last_updated": None},
        "NHS456": {"nhs_number": "NHS456", "created": None, "last_updated": None},
    }

    with pytest.raises(OdsReportException):
        ods_report_service.create_and_save_ods_report(
            ods_code,
            patient_rows,
            upload_to_s3=True,
            create_pre_signed_url=True,
            file_type_output="invalid",
        )

    mock_create_report_pdf.assert_not_called()
    mock_save_report_to_s3.assert_not_called()
    mock_get_pre_signed_url.assert_not_called()


@freeze_time("2024-01-01T12:00:00Z")
def test_create_and_save_ods_report_with_pre_sign_url(
    ods_report_service,
    mock_create_report_csv,
    mock_save_report_to_s3,
    mock_get_pre_signed_url,
    mocker,
):
    ods_code = "ODS123"
    patient_rows = {
        "NHS123": {"nhs_number": "NHS123", "created": None, "last_updated": None},
        "NHS456": {"nhs_number": "NHS456", "created": None, "last_updated": None},
    }
    file_name = "LloydGeorgeSummary_ODS123_2_2024-01-01_12-00.csv"
    mock_pre_sign_url = "https://presigned.url"
    temp_file_path = os.path.join(ods_report_service.temp_output_dir, file_name)

    mock_get_pre_signed_url.return_value = mock_pre_sign_url
    mocker.patch.object(
        ods_report_service,
        "create_ods_report",
        return_value=(file_name, temp_file_path),
    )

    result = ods_report_service.create_and_save_ods_report(
        ods_code,
        patient_rows,
        True,
        True,
        FileType.CSV,
    )

    mock_save_report_to_s3.assert_called_once_with(ods_code, file_name, temp_file_path)
    mock_get_pre_signed_url.assert_called_once_with(ods_code, file_name)
    assert result == mock_pre_sign_url


def test_create_report_csv(ods_report_service, tmp_path):
    patient_rows = {
        "NHS123": {"nhs_number": "NHS123", "created": None, "last_updated": None},
        "NHS456": {"nhs_number": "NHS456", "created": None, "last_updated": None},
    }
    file_name = tmp_path / "test_report.csv"
    ods_code = "ODS123"

    ods_report_service.create_csv_report(str(file_name), patient_rows, ods_code)

    with open(file_name, "r") as f:
        content = f.readlines()

    assert (
        f"Total number of patients for ODS code {ods_code}: {len(patient_rows)}\n"
        in content
    )
    assert "NHS Number,Latest Created Date,Latest Updated Date\n" in content


def test_create_xlsx_report(ods_report_service, tmp_path):
    file_name = tmp_path / "test_report.xlsx"
    patient_rows = {
        "NHS123456": {"nhs_number": "NHS123456", "created": None, "last_updated": None},
        "NHS654321": {"nhs_number": "NHS654321", "created": None, "last_updated": None},
        "NHS111222": {"nhs_number": "NHS111222", "created": None, "last_updated": None},
    }
    ods_code = "ODS123"

    ods_report_service.create_xlsx_report(str(file_name), patient_rows, ods_code)

    assert os.path.exists(file_name)

    wb = load_workbook(str(file_name))
    ws = wb.active

    assert (
        ws["A1"].value
        == f"Total number of patients for ODS code {ods_code}: {len(patient_rows)}\n"
    )

    nhs_sorted = sorted(patient_rows.keys())
    for i, nhs_number in enumerate(nhs_sorted, start=3):
        assert ws.cell(row=i, column=1).value == nhs_number


@freeze_time("2024-01-01T12:00:00Z")
def test_create_pdf_report(ods_report_service, tmp_path):
    file_name = tmp_path / "test_report.pdf"
    patient_rows = {
        "NHS123456": {"nhs_number": "NHS123456", "created": None, "last_updated": None},
        "NHS654321": {"nhs_number": "NHS654321", "created": None, "last_updated": None},
        "NHS111222": {"nhs_number": "NHS111222", "created": None, "last_updated": None},
    }
    ods_code = "ODS123"

    ods_report_service.create_pdf_report(str(file_name), patient_rows, ods_code)

    assert os.path.exists(file_name)

    reader = PdfReader(str(file_name))
    assert len(reader.pages) > 0

    first_page = reader.pages[0].extract_text() or ""

    assert f"NHS numbers within NDR for ODS code: {ods_code}" in first_page
    assert f"Total number of patients: {len(patient_rows)}" in first_page
    for nhs_number in patient_rows.keys():
        assert nhs_number in first_page


@freeze_time("2024-01-01T12:00:00Z")
def test_save_report_to_s3(ods_report_service, mocker):
    ods_report_service.s3_service = mocker.Mock()
    mocker.patch.object(ods_report_service.s3_service, "upload_file")

    ods_report_service.save_report_to_s3("ODS123", "test_report.csv", "path/to/file")

    ods_report_service.s3_service.upload_file.assert_called_once_with(
        s3_bucket_name="test_statistics_report_bucket",
        file_key="ods-reports/ODS123/2024/1/1/test_report.csv",
        file_name="path/to/file",
    )


@freeze_time("2024-01-01T12:00:00Z")
def test_get_pre_signed_url(ods_report_service, mocker):
    ods_report_service.s3_service = mocker.Mock()
    mocker.patch.object(ods_report_service.s3_service, "create_download_presigned_url")

    ods_report_service.get_pre_signed_url("ODS123", "test_report.csv")

    ods_report_service.s3_service.create_download_presigned_url.assert_called_once_with(
        s3_bucket_name="test_statistics_report_bucket",
        file_key="ods-reports/ODS123/2024/1/1/test_report.csv",
    )


def test_get_documents_for_review(
    ods_report_service,
    mock_get_pre_signed_url,
    mock_review_result,
    mock_save_report_to_s3,
    mock_create_review_csv_report,
    mocker,
):
    expected_result = "https://example.com/mocked"
    mock_get_pre_signed_url.return_value = expected_result
    mock_ods_code = "ODS123"

    query_builder = DynamoQueryFilterBuilder()
    query_builder.add_condition(
        "ReviewStatus",
        AttributeOperator.EQUAL,
        DocumentReviewStatus.PENDING_REVIEW,
    )
    expected_query_filter = query_builder.build()

    mocker.patch.object(
        ods_report_service.document_upload_review_service,
        "fetch_documents_from_table",
        return_value=[mock_review_result, mock_review_result, mock_review_result],
    )

    mocker.patch.object(
        ods_report_service.document_upload_review_service,
        "build_review_dynamo_filter",
        return_value=expected_query_filter,
    )

    result = ods_report_service.get_documents_for_review(mock_ods_code, FileType.CSV)

    assert result == expected_result

    ods_report_service.document_upload_review_service.fetch_documents_from_table.assert_called_once_with(
        search_key="Custodian",
        search_condition=mock_ods_code,
        index_name="CustodianIndex",
        query_filter=expected_query_filter,
    )

    mock_save_report_to_s3.assert_called_once()
    mock_create_review_csv_report.assert_called_once()


def test_get_documents_for_review_unsupported_file_type(ods_report_service):
    with pytest.raises(OdsReportException) as excinfo:
        ods_report_service.get_documents_for_review("mock_ods_code", FileType.PDF)

    assert excinfo.value.error == LambdaError.UnsupportedFileType
    assert excinfo.value.status_code == 400


@freeze_time("2024-01-01T12:00:00Z")
def test_create_review_csv_report(ods_report_service, mock_review_result, tmp_path):
    file_name = tmp_path / "test_review_report.csv"
    data = [mock_review_result, mock_review_result, mock_review_result]

    ods_report_service.create_review_csv_report(str(file_name), data)

    assert os.path.exists(file_name)

    with open(file_name, "r") as f:
        content = f.readlines()

    assert (
        content[0]
        == "nhs_number,review_reason,document_snomed_code_type,author,upload_date\n"
    )


def test_query_table_by_index(ods_report_service, mocked_context, set_env):
    mock_dynamo_results = [{"mock_database_field": "mock_database_value"}]
    mock_ods_code = "ODS123"

    ods_report_service.dynamo_service.query_table.return_value = mock_dynamo_results

    response = ods_report_service.query_table_by_index(mock_ods_code)

    assert response == mock_dynamo_results

    ods_report_service.dynamo_service.query_table.assert_called_once_with(
        table_name=os.getenv("LLOYD_GEORGE_DYNAMODB_NAME"),
        index_name="OdsCodeIndex",
        search_key=DocumentReferenceMetadataFields.CURRENT_GP_ODS.value,
        search_condition=mock_ods_code,
        query_filter=FinalStatusAndNotSuperseded,
    )


def test_query_table_by_index_pcse_user(
    ods_report_service,
    mocked_pcse_context,
    set_env,
):
    mock_dynamo_results = [{"mock_database_field": "mock_database_value"}]
    mock_ods_code = "ODS123"

    expected_result = mock_dynamo_results + mock_dynamo_results

    ods_report_service.dynamo_service.query_table.return_value = mock_dynamo_results

    response = ods_report_service.query_table_by_index(mock_ods_code)

    assert response == expected_result

    ods_report_service.dynamo_service.query_table.assert_has_calls(
        [
            call(
                table_name=os.getenv("LLOYD_GEORGE_DYNAMODB_NAME"),
                index_name="OdsCodeIndex",
                search_key=DocumentReferenceMetadataFields.CURRENT_GP_ODS.value,
                search_condition=PatientOdsInactiveStatus.SUSPENDED,
                query_filter=FinalStatusAndNotSuperseded,
            ),
            call(
                table_name=os.getenv("LLOYD_GEORGE_DYNAMODB_NAME"),
                index_name="OdsCodeIndex",
                search_key=DocumentReferenceMetadataFields.CURRENT_GP_ODS.value,
                search_condition=PatientOdsInactiveStatus.DECEASED,
                query_filter=FinalStatusAndNotSuperseded,
            ),
        ],
    )


def test_query_table_by_index_no_results_errors(ods_report_service, mocked_context):
    mock_ods_code = "ODS123"

    ods_report_service.dynamo_service.query_table.return_value = []

    with pytest.raises(OdsReportException) as excinfo:
        ods_report_service.query_table_by_index(mock_ods_code)

    assert excinfo.value.status_code == 404
    assert excinfo.value.error == LambdaError.NoDataFound


def test_create_and_save_ods_report_dont_save_to_s3(
    ods_report_service,
    mock_save_report_to_s3,
    mock_get_pre_signed_url,
    mocker,
):
    mocker.patch.object(
        ods_report_service,
        "create_ods_report",
        return_value=("x.csv", "/tmp/x.csv"),
    )

    ods_report_service.create_and_save_ods_report(
        ods_code="ODS123",
        patient_rows={
            "mock_nhs_number": {
                "nhs_number": "mock_nhs_number",
                "created": None,
                "last_updated": None,
            },
        },
        create_pre_signed_url=False,
        upload_to_s3=False,
        file_type_output=FileType.CSV,
    )

    mock_save_report_to_s3.assert_not_called()
    mock_get_pre_signed_url.assert_not_called()


def test_create_pdf_report_multiple_pages(ods_report_service, tmp_path):
    file_name = tmp_path / "test_report.pdf"
    patient_rows = {}

    for i in range(1, 150):
        nhs = f"NHS123456{str(i)}"
        patient_rows[nhs] = {"nhs_number": nhs, "created": None, "last_updated": None}

    ods_report_service.create_pdf_report(str(file_name), patient_rows, "ODS123")

    assert os.path.exists(file_name)

    reader = PdfReader(str(file_name))
    assert len(reader.pages) > 1


def test_build_patient_rows_skips_items_without_nhs_number(ods_report_service):
    items = [
        {
            DocumentReferenceMetadataFields.CREATED.value: "2026-02-25T12:50:35.000000Z",
            DocumentReferenceMetadataFields.LAST_UPDATED.value: 1700000000,
        },
        {
            DocumentReferenceMetadataFields.NHS_NUMBER.value: "9730786917",
            DocumentReferenceMetadataFields.CREATED.value: "2026-02-25T12:50:35.000000Z",
            DocumentReferenceMetadataFields.LAST_UPDATED.value: 1700000001,
        },
    ]

    rows = ods_report_service.build_patient_rows(items)

    assert list(rows.keys()) == ["9730786917"]
    assert rows["9730786917"]["nhs_number"] == "9730786917"


def test_build_patient_rows_dedupes_and_picks_earliest_created_and_latest_last_updated(
    ods_report_service,
):
    nhs = "9730786917"

    items = [
        {
            DocumentReferenceMetadataFields.NHS_NUMBER.value: nhs,
            DocumentReferenceMetadataFields.CREATED.value: "2026-02-25T12:50:40.000000Z",
            DocumentReferenceMetadataFields.LAST_UPDATED.value: 1700000010,
        },
        {
            DocumentReferenceMetadataFields.NHS_NUMBER.value: nhs,
            DocumentReferenceMetadataFields.CREATED.value: "2026-02-25T12:50:35.000000Z",
            DocumentReferenceMetadataFields.LAST_UPDATED.value: 1700000020,
        },
        {
            DocumentReferenceMetadataFields.NHS_NUMBER.value: "9730786933",
            DocumentReferenceMetadataFields.CREATED.value: "2026-02-25T12:50:36.000000Z",
            DocumentReferenceMetadataFields.LAST_UPDATED.value: 1700000030,
        },
    ]

    rows = ods_report_service.build_patient_rows(items)

    assert set(rows.keys()) == {nhs, "9730786933"}

    expected_created = datetime(2026, 2, 25, 12, 50, 40, tzinfo=timezone.utc)
    expected_last_updated = datetime.fromtimestamp(1700000020, tz=timezone.utc)

    assert rows[nhs]["latest_created_date"] == expected_created
    assert rows[nhs]["latest_updated_date"] == expected_last_updated


def test_build_patient_rows_does_not_overwrite_existing_values_with_none(
    ods_report_service,
):
    nhs = "9730786917"

    items = [
        {
            DocumentReferenceMetadataFields.NHS_NUMBER.value: nhs,
            DocumentReferenceMetadataFields.CREATED.value: "2026-02-25T12:50:35.000000Z",
            DocumentReferenceMetadataFields.LAST_UPDATED.value: 1700000010,
        },
        {
            DocumentReferenceMetadataFields.NHS_NUMBER.value: nhs,
            DocumentReferenceMetadataFields.CREATED.value: None,
            DocumentReferenceMetadataFields.LAST_UPDATED.value: None,
        },
    ]

    rows = ods_report_service.build_patient_rows(items)

    expected_created = datetime(2026, 2, 25, 12, 50, 35, tzinfo=timezone.utc)
    expected_last_updated = datetime.fromtimestamp(1700000010, tz=timezone.utc)

    assert rows[nhs]["latest_created_date"] == expected_created
    assert rows[nhs]["latest_updated_date"] == expected_last_updated
