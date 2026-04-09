from types import SimpleNamespace

import pytest
from freezegun import freeze_time
from openpyxl import load_workbook

from services.reporting.excel_report_generator_service import ExcelReportGenerator


@pytest.fixture
def excel_report_generator():
    return ExcelReportGenerator()


@pytest.fixture
def mock_ods_report(mocker):
    report = mocker.Mock(name="OdsReport")
    report.get_total_ingested_count.return_value = 2
    report.get_total_successful.return_value = 1
    report.get_total_in_review_count.return_value = 1
    report.get_total_in_review_percentage.return_value = "50%"
    report.get_total_successful_percentage.return_value = "50%"
    report.get_total_registered_elsewhere_count.return_value = 0
    report.get_total_suspended_count.return_value = 0
    report.get_total_deceased_count.return_value = 0
    report.get_total_restricted_count.return_value = 0
    report.get_unsuccessful_reasons_data_rows.return_value = [
        ["Reason", "Invalid NHS number", 1],
    ]
    return report


@pytest.fixture
def make_record():
    def _make(
        *,
        nhs_number=None,
        date=None,
        uploader_ods_code=None,
        pds_ods_code=None,
        upload_status=None,
        reason=None,
        sent_to_review=None,
        file_path=None,
    ):
        return SimpleNamespace(
            nhs_number=nhs_number,
            date=date,
            uploader_ods_code=uploader_ods_code,
            pds_ods_code=pds_ods_code,
            upload_status=upload_status,
            reason=reason,
            sent_to_review=sent_to_review,
            file_path=file_path,
        )

    return _make


@pytest.fixture
def make_report(tmp_path, excel_report_generator, mock_ods_report, mocker):
    def _make(*, ods_code="Y12345", records=None, filename="report.xlsx"):
        records = records or []
        output_file = tmp_path / filename

        mocked_build = mocker.patch.object(
            excel_report_generator,
            "_build_ods_report",
            return_value=mock_ods_report,
        )

        result = excel_report_generator.create_report_orchestration_xlsx(
            ods_code=ods_code,
            records=records,
            output_path=str(output_file),
        )

        assert result == str(output_file)
        assert output_file.exists()

        wb = load_workbook(output_file)
        detail_ws = wb["Daily Upload Report"]
        summary_ws = wb["Summary"]

        mocked_build.assert_called_once()

        return output_file, detail_ws, summary_ws

    return _make


@pytest.fixture
def expected_header_row():
    return [
        "NHS Number",
        "Date",
        "Uploader ODS",
        "PDS ODS",
        "Upload Status",
        "Reason",
        "Sent to Review",
        "File Path",
    ]


@freeze_time("2025-01-01T12:00:00")
def test_create_report_orchestration_xlsx_happy_path(
    make_report,
    make_record,
    expected_header_row,
):
    ods_code = "Y12345"
    records = [
        make_record(
            nhs_number="1234567890",
            date="2025-01-01",
            uploader_ods_code="Y12345",
            pds_ods_code="A99999",
            upload_status="SUCCESS",
            reason=None,
            sent_to_review=False,
            file_path="/path/file1.pdf",
        ),
        make_record(
            nhs_number="123456789",
            date="2025-01-02",
            uploader_ods_code="Y12345",
            pds_ods_code="B88888",
            upload_status="FAILED",
            reason="Invalid NHS number",
            sent_to_review=True,
            file_path="/path/file2.pdf",
        ),
    ]

    _, ws, summary_ws = make_report(
        ods_code=ods_code,
        records=records,
        filename="report.xlsx",
    )

    assert ws.title == "Daily Upload Report"
    assert ws["A1"].value == f"ODS Code: {ods_code}"
    assert ws["A2"].value == "Generated at (UTC): 2025-01-01T12:00:00+00:00"
    assert ws["A3"].value is None

    assert [cell.value for cell in ws[4]] == expected_header_row

    assert [cell.value for cell in ws[5]] == [
        "1234567890",
        "2025-01-01",
        "Y12345",
        "A99999",
        "SUCCESS",
        None,
        False,
        "/path/file1.pdf",
    ]

    assert [cell.value for cell in ws[6]] == [
        "123456789",
        "2025-01-02",
        "Y12345",
        "B88888",
        "FAILED",
        "Invalid NHS number",
        True,
        "/path/file2.pdf",
    ]

    assert summary_ws.title == "Summary"
    assert [cell.value for cell in summary_ws[1]] == ["Type", "Description", "Count"]


def test_create_report_orchestration_xlsx_with_no_records(
    make_report,
    expected_header_row,
):
    _, ws, summary_ws = make_report(records=[], filename="empty_report.xlsx")

    assert ws.max_row == 4
    assert [cell.value for cell in ws[4]] == expected_header_row
    assert [cell.value for cell in summary_ws[1]] == ["Type", "Description", "Count"]


@pytest.mark.parametrize(
    "records, expected_row",
    [
        (
            [
                SimpleNamespace(
                    nhs_number="1234567890",
                    date=None,
                    uploader_ods_code=None,
                    pds_ods_code=None,
                    upload_status=None,
                    reason=None,
                    sent_to_review=None,
                    file_path=None,
                ),
            ],
            ["1234567890", None, None, None, None, None, None, None],
        ),
        (
            [
                SimpleNamespace(
                    nhs_number=None,
                    date="2025-01-01",
                    uploader_ods_code=None,
                    pds_ods_code=None,
                    upload_status=None,
                    reason=None,
                    sent_to_review=None,
                    file_path=None,
                ),
            ],
            [None, "2025-01-01", None, None, None, None, None, None],
        ),
        (
            [
                SimpleNamespace(
                    nhs_number=None,
                    date=None,
                    uploader_ods_code="Y12345",
                    pds_ods_code=None,
                    upload_status="SUCCESS",
                    reason=None,
                    sent_to_review=None,
                    file_path=None,
                ),
            ],
            [None, None, "Y12345", None, "SUCCESS", None, None, None],
        ),
    ],
)
def test_create_report_orchestration_xlsx_handles_missing_fields(
    make_report,
    expected_header_row,
    records,
    expected_row,
):
    _, ws, _summary_ws = make_report(records=records, filename="partial.xlsx")

    assert [cell.value for cell in ws[4]] == expected_header_row
    assert [cell.value for cell in ws[5]] == expected_row
